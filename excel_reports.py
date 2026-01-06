# -*- coding: utf-8 -*-
"""
Reportes Excel - Sistema de Mensualidades
==========================================

Genera reportes Excel profesionales con múltiples hojas:

REPORTES DISPONIBLES:
1) 📊 Reporte Completo (Multi-hoja):
   - Resumen ejecutivo (KPIs)
   - Estudiantes (control académico/financiero)
   - Pagos (auditoría/contabilidad)
   - Morosos y Vencidos (cobranza)
   - Cursos (análisis por curso)

2) 👥 Reporte de Estudiantes (compatible con ruta antigua)
3) 💳 Reporte de Pagos (compatible con ruta antigua)
4) 🚨 Reporte de Próximos a Vencer (compatible con ruta antigua)

COMPATIBILIDAD 100% con tu sistema actual:
- Cliente: id, nombre, apellido, nombre_completo, cedula, email, telefono, 
           activo, fecha_registro, fecha_inicio_clases, fecha_fin,
           dias_restantes, mensualidades_canceladas, valor_inscripcion,
           curso (relación), estado_pago (propiedad), total_pagado, saldo_pendiente
- Curso: nombre, precio_mensual, precio_inscripcion, duracion_meses, activo
- Pago: id, fecha_pago, monto, metodo_pago, referencia, periodo, notas, cliente
"""

import io
from datetime import datetime
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _money(v):
    """Convierte a float de forma segura"""
    try:
        return float(v or 0)
    except Exception:
        return 0.0


class ExcelReportGenerator:
    """Generador de reportes Excel para el sistema de mensualidades"""
    
    def __init__(self):
        self.workbook = None
        
        # 🎨 PALETA DE COLORES (sin #)
        self.colors = {
            "header": "1F4E78",        # Azul oscuro
            "subheader": "2F75B5",     # Azul medio
            "ok": "70AD47",            # Verde (al día)
            "warn": "FFC000",          # Amarillo (por vencer)
            "danger": "C00000",        # Rojo (vencido)
            "soft": "F2F2F2",          # Gris suave
            "total": "FFD966",         # Amarillo suave (totales)
            "info": "5B9BD5",          # Azul info
        }
        
        # Bordes
        thin = Side(style="thin", color="BFBFBF")
        self.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        
        # Fuentes
        self.font_title = Font(bold=True, size=18, color="FFFFFF")
        self.font_header = Font(bold=True, size=11, color="FFFFFF")
        self.font_bold = Font(bold=True)
        self.font_normal = Font(size=10)
        
        # Alineaciones
        self.align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.align_right = Alignment(horizontal="right", vertical="center")
        
        # Rellenos
        self.fill_header = PatternFill(
            start_color=self.colors["header"],
            end_color=self.colors["header"],
            fill_type="solid"
        )
        self.fill_subheader = PatternFill(
            start_color=self.colors["subheader"],
            end_color=self.colors["subheader"],
            fill_type="solid"
        )
        self.fill_soft = PatternFill(
            start_color=self.colors["soft"],
            end_color=self.colors["soft"],
            fill_type="solid"
        )
    
    # ============================================
    # 🔧 UTILIDADES DE FORMATO
    # ============================================
    
    def _set_col_widths_auto(self, ws, min_width=10, max_width=50, padding=2):
        """Auto-ajusta el ancho de columnas según contenido"""
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                if cell.value is None:
                    continue
                s = str(cell.value)
                if len(s) > max_len:
                    max_len = len(s)
            width = max(min_width, min(max_width, max_len + padding))
            ws.column_dimensions[col_letter].width = width
    
    def _apply_table_header(self, ws, row, headers):
        """Aplica formato de encabezado a una fila"""
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = self.font_header
            cell.fill = self.fill_subheader
            cell.alignment = self.align_center
            cell.border = self.border
        ws.row_dimensions[row].height = 22
    
    def _title_block(self, ws, title, subtitle=None, merge_to_col=8):
        """Crea bloque de título en las primeras dos filas"""
        end_col = get_column_letter(merge_to_col)
        
        # Fila 1: Título
        ws.merge_cells(f"A1:{end_col}1")
        t = ws["A1"]
        t.value = title
        t.font = self.font_title
        t.fill = self.fill_header
        t.alignment = self.align_center
        ws.row_dimensions[1].height = 36
        
        # Fila 2: Subtítulo o fecha
        ws.merge_cells(f"A2:{end_col}2")
        s = ws["A2"]
        s.value = subtitle or f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        s.font = Font(italic=True, size=10, color="404040")
        s.alignment = self.align_center
        ws.row_dimensions[2].height = 20
    
    def _currency(self, cell):
        """Aplica formato de moneda a una celda"""
        cell.number_format = '"$"#,##0.00'
    
    def _date_format(self, cell):
        """Aplica formato de fecha a una celda"""
        cell.number_format = 'DD/MM/YYYY'
    
    # ============================================
    # 📊 CÁLCULOS Y CLASIFICACIONES
    # ============================================
    
    def _clasificar_estudiantes(self, estudiantes):
        """Clasifica estudiantes según su estado de pago"""
        activos = [e for e in estudiantes if getattr(e, "activo", True)]
        
        sin_cobertura = []
        vencidos = []
        criticos = []      # 0-3 días
        proximos = []      # 4-7 días
        al_dia = []        # >7 días
        
        for e in activos:
            # Sin cobertura: no ha pagado ninguna mensualidad
            if getattr(e, "mensualidades_canceladas", 0) == 0:
                sin_cobertura.append(e)
                continue
            
            dias = getattr(e, "dias_restantes", None)
            
            if dias is None:
                sin_cobertura.append(e)
            elif dias < 0:
                vencidos.append(e)
            elif dias <= 3:
                criticos.append(e)
            elif dias <= 7:
                proximos.append(e)
            else:
                al_dia.append(e)
        
        return activos, sin_cobertura, vencidos, criticos, proximos, al_dia
    
    def _pagos_por_cliente(self, pagos):
        """Agrupa pagos por cliente_id"""
        mp = defaultdict(list)
        for p in pagos:
            cid = getattr(p, "cliente_id", None)
            if cid is None:
                c = getattr(p, "cliente", None)
                if c:
                    cid = getattr(c, "id", None)
            if cid:
                mp[cid].append(p)
        return mp
    
    # ============================================
    # 📄 REPORTE COMPLETO (MULTI-HOJA)
    # ============================================
    
    def generar_reporte_completo(self, estudiantes, pagos, cursos):
        """
        Genera Excel multi-hoja con análisis completo del sistema
        
        Args:
            estudiantes: Lista de objetos Cliente
            pagos: Lista de objetos Pago
            cursos: Lista de objetos Curso
        
        Returns:
            BytesIO con el archivo Excel
        """
        self.workbook = openpyxl.Workbook()
        
        # Quitar hoja por defecto
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        # Crear hojas en orden
        self._sheet_resumen(estudiantes, pagos, cursos)
        self._sheet_estudiantes(estudiantes, pagos)
        self._sheet_pagos(pagos)
        self._sheet_morosos(estudiantes, pagos)
        self._sheet_cursos(estudiantes, pagos, cursos)
        
        # Guardar en memoria
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output
    
    # ============================================
    # 📊 HOJA 1: RESUMEN EJECUTIVO
    # ============================================
    
    def _sheet_resumen(self, estudiantes, pagos, cursos):
        """Hoja con KPIs y métricas ejecutivas"""
        ws = self.workbook.create_sheet("📊 Resumen", 0)
        self._title_block(ws, "RESUMEN EJECUTIVO", merge_to_col=4)
        
        # Clasificar estudiantes
        activos, sin_cobertura, vencidos, criticos, proximos, al_dia = self._clasificar_estudiantes(estudiantes)
        
        # Cálculos financieros
        total_recaudado = sum(_money(p.monto) for p in pagos)
        total_pagos = len(pagos)
        promedio = (total_recaudado / total_pagos) if total_pagos else 0
        
        # Ingresos del mes actual
        now = datetime.now()
        inicio_mes = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        pagos_mes = [p for p in pagos if getattr(p, "fecha_pago", now) >= inicio_mes]
        total_mes = sum(_money(p.monto) for p in pagos_mes)
        
        # Cursos activos
        cursos_activos = [c for c in cursos if getattr(c, "activo", True)]
        
        # ===================================
        # TABLA DE KPIs
        # ===================================
        ws["A3"].value = "INDICADORES CLAVE"
        ws["A3"].font = Font(bold=True, size=12)
        ws.merge_cells("A3:D3")
        ws["A3"].fill = self.fill_soft
        ws["A3"].alignment = self.align_left
        
        kpis = [
            ("Total estudiantes activos", len(activos), None),
            ("Sin cobertura (no han pagado)", len(sin_cobertura), "danger" if len(sin_cobertura) else "ok"),
            ("Vencidos", len(vencidos), "danger" if len(vencidos) else "ok"),
            ("Críticos (≤3 días)", len(criticos), "warn" if len(criticos) else "ok"),
            ("Próximos (4-7 días)", len(proximos), "warn" if len(proximos) else "ok"),
            ("Al día (>7 días)", len(al_dia), "ok"),
            ("", "", None),  # Separador
            ("Total recaudado (histórico)", total_recaudado, "ok"),
            ("Total recaudado (este mes)", total_mes, "ok"),
            ("Cantidad de pagos", total_pagos, None),
            ("Promedio por pago", promedio, None),
            ("", "", None),  # Separador
            ("Cursos activos", len(cursos_activos), None),
        ]
        
        row = 4
        for label, value, tag in kpis:
            if label == "":
                row += 1
                continue
            
            # Columna A: Etiqueta
            ws.cell(row=row, column=1, value=label).alignment = self.align_left
            ws.cell(row=row, column=1).border = self.border
            ws.cell(row=row, column=1).font = self.font_normal
            
            # Columna B: Valor
            c2 = ws.cell(row=row, column=2, value=value)
            c2.border = self.border
            c2.alignment = self.align_center
            c2.font = Font(bold=True, size=12)
            
            # Formato moneda si aplica
            if "recaudado" in label.lower() or "promedio" in label.lower():
                self._currency(c2)
            
            # Color según severidad
            if tag == "danger":
                c2.fill = PatternFill(
                    start_color=self.colors["danger"],
                    end_color=self.colors["danger"],
                    fill_type="solid"
                )
                c2.font = Font(bold=True, size=12, color="FFFFFF")
            elif tag == "warn":
                c2.fill = PatternFill(
                    start_color=self.colors["warn"],
                    end_color=self.colors["warn"],
                    fill_type="solid"
                )
                c2.font = Font(bold=True, size=12)
            elif tag == "ok":
                c2.fill = PatternFill(
                    start_color=self.colors["ok"],
                    end_color=self.colors["ok"],
                    fill_type="solid"
                )
                c2.font = Font(bold=True, size=12, color="FFFFFF")
            
            row += 1
        
        # Ajustar anchos
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 5
        ws.column_dimensions["D"].width = 5
        
        ws.freeze_panes = "A4"
    
    # ============================================
    # 👥 HOJA 2: ESTUDIANTES
    # ============================================
    

    def _sheet_estudiantes(self, estudiantes, pagos):
        """Hoja con detalle de todos los estudiantes"""
        ws = self.workbook.create_sheet("👥 Estudiantes")
        self._title_block(ws, "REGISTRO DE ESTUDIANTES", merge_to_col=16)
    
        pagos_por_cliente = self._pagos_por_cliente(pagos)
    
        headers = [
            "ID", "Nombre Completo", "Cédula", "Email", "Teléfono", "Curso",
            "Precio/Mes", "F. Registro", "Inicio Clases",
            "F. Vencimiento", "Días Rest.",
            "Mens. Pagadas", "Inscripción",
            "Total Pagado", "Estado", "Activo"
        ]
        self._apply_table_header(ws, 4, headers)
    
        row = 5
    
        # ✅ VARIABLES PARA TOTALES
        total_inscripciones = 0.0
        total_pagado_general = 0.0
    
        for e in estudiantes:
            curso = getattr(e, "curso", None)
            precio_mensual = _money(getattr(curso, "precio_mensual", 0)) if curso else 0
        
            # Total pagado por este estudiante
            pagos_cliente = pagos_por_cliente.get(getattr(e, "id", None), [])
            total_pagado = sum(_money(p.monto) for p in pagos_cliente)
        
            # ✅ Acumular totales
            inscripcion_valor = _money(getattr(e, "abono_inscripcion", 0))
            total_inscripciones += inscripcion_valor
            total_pagado_general += total_pagado
        
            # Determinar estado
            dias = getattr(e, "dias_restantes", None)
            mens_canceladas = getattr(e, "mensualidades_canceladas", 0)
        
            if mens_canceladas == 0:
                estado = "Sin cobertura"
                sem = "danger"
            elif dias is None:
                estado = "N/A"
                sem = None
            elif dias < 0:
                estado = "Vencido"
                sem = "danger"
            elif dias <= 7:
                estado = "Por vencer"
                sem = "warn"
            else:
                estado = "Al día"
                sem = "ok"
        
            # Construir nombre completo
            nombre_completo = getattr(e, "nombre_completo", "")
            if not nombre_completo:
                nombre = getattr(e, "nombre", "")
                apellido = getattr(e, "apellido", "")
                nombre_completo = f"{nombre} {apellido}".strip()
        
            values = [
                getattr(e, "id", ""),
                nombre_completo,
                getattr(e, "cedula", None) or "Sin registrar",
                getattr(e, "email", ""),
                getattr(e, "telefono", None) or "N/A",
                getattr(curso, "nombre", None) if curso else "Sin curso",
                precio_mensual,
                getattr(e, "fecha_registro", None),
                getattr(e, "fecha_inicio_clases", None),
                getattr(e, "fecha_fin", None),
                dias if dias is not None else "N/A",
                mens_canceladas,
                inscripcion_valor,  # ✅ Usar la variable que acumulamos
                total_pagado,
                estado,
                "Sí" if getattr(e, "activo", True) else "No",
            ]
        
            for c, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.border = self.border
                cell.alignment = self.align_left
                cell.font = self.font_normal
            
                # Alineaciones especiales
                if c in (1, 11, 12, 16):  # ID, días, mensualidades, activo
                    cell.alignment = self.align_center
                if c in (7, 13, 14):  # Montos
                    cell.alignment = self.align_right
                    self._currency(cell)
            
                # Fechas
                if c in (8, 9, 10) and v and v != "N/A":
                    if isinstance(v, datetime):
                        cell.value = v.strftime("%d/%m/%Y")
                    cell.alignment = self.align_center
            
                # Colorear estado (columna 15)
                if c == 15 and sem:
                    if sem == "danger":
                        cell.fill = PatternFill(
                            start_color=self.colors["danger"],
                            end_color=self.colors["danger"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.alignment = self.align_center
                    elif sem == "warn":
                        cell.fill = PatternFill(
                            start_color=self.colors["warn"],
                            end_color=self.colors["warn"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True)
                        cell.alignment = self.align_center
                    elif sem == "ok":
                        cell.fill = PatternFill(
                            start_color=self.colors["ok"],
                            end_color=self.colors["ok"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.alignment = self.align_center
        
            row += 1
    
        # ✅ AGREGAR FILA DE TOTALES
        row += 1  # Dejar una fila en blanco
    
        # Merge las primeras columnas para el texto "TOTALES"
        ws.merge_cells(f"A{row}:L{row}")
        total_cell = ws.cell(row=row, column=1, value="TOTALES")
        total_cell.font = Font(bold=True, size=13)
        total_cell.fill = PatternFill(
            start_color=self.colors["total"],
            end_color=self.colors["total"],
            fill_type="solid"
        )
        total_cell.alignment = self.align_center
        total_cell.border = self.border
    
        # Columna M (13): Total Inscripciones
        inscripcion_cell = ws.cell(row=row, column=13, value=total_inscripciones)
        inscripcion_cell.font = Font(bold=True, size=13)
        inscripcion_cell.fill = PatternFill(
            start_color=self.colors["total"],
            end_color=self.colors["total"],
            fill_type="solid"
        )
        inscripcion_cell.alignment = self.align_right
        inscripcion_cell.border = self.border
        self._currency(inscripcion_cell)
    
        # Columna N (14): Total Pagado
        pagado_cell = ws.cell(row=row, column=14, value=total_pagado_general)
        pagado_cell.font = Font(bold=True, size=13)
        pagado_cell.fill = PatternFill(
            start_color=self.colors["total"],
            end_color=self.colors["total"],
            fill_type="solid"
        )
        pagado_cell.alignment = self.align_right
        pagado_cell.border = self.border
        self._currency(pagado_cell)
    
        # Columnas vacías (15-16)
        for col in [15, 16]:
            empty_cell = ws.cell(row=row, column=col)
            empty_cell.fill = PatternFill(
                start_color=self.colors["total"],
                end_color=self.colors["total"],
                fill_type="solid"
            )
            empty_cell.border = self.border
    
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, row-2)}"
        self._set_col_widths_auto(ws)
    
        # ============================================
        # 💳 HOJA 3: PAGOS
        # ============================================
    
    def _sheet_pagos(self, pagos):
        """Hoja con historial completo de pagos"""
        ws = self.workbook.create_sheet("💳 Pagos")
        self._title_block(ws, "HISTORIAL DE PAGOS", merge_to_col=10)
        
        headers = [
            "ID", "Fecha", "Estudiante", "Cédula", "Curso",
            "Periodo", "Monto", "Método", "Referencia", "Notas"
        ]
        self._apply_table_header(ws, 4, headers)
        
        row = 5
        total = 0.0
        
        for p in pagos:
            c = getattr(p, "cliente", None)
            curso = getattr(c, "curso", None) if c else None
            
            monto = _money(getattr(p, "monto", 0))
            total += monto
            
            # Nombre completo del cliente
            if c:
                nombre_completo = getattr(c, "nombre_completo", "")
                if not nombre_completo:
                    nombre = getattr(c, "nombre", "")
                    apellido = getattr(c, "apellido", "")
                    nombre_completo = f"{nombre} {apellido}".strip()
            else:
                nombre_completo = "N/A"
            
            values = [
                getattr(p, "id", ""),
                getattr(p, "fecha_pago", None),
                nombre_completo,
                getattr(c, "cedula", None) if c else None,
                getattr(curso, "nombre", None) if curso else "Sin curso",
                getattr(p, "periodo", None) or "-",
                monto,
                getattr(p, "metodo_pago", None) or "-",
                getattr(p, "referencia", None) or "-",
                getattr(p, "notas", None) or "-",
            ]
            
            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = self.border
                cell.alignment = self.align_left
                cell.font = self.font_normal
                
                if col == 1:  # ID
                    cell.alignment = self.align_center
                if col == 2 and v:  # Fecha
                    if isinstance(v, datetime):
                        cell.value = v.strftime("%d/%m/%Y %H:%M")
                    cell.alignment = self.align_center
                if col == 7:  # Monto
                    cell.alignment = self.align_right
                    self._currency(cell)
            
            row += 1
        
        # Fila de total
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        tcell = ws.cell(row=row, column=1, value="TOTAL RECAUDADO")
        tcell.font = Font(bold=True, size=13)
        tcell.fill = PatternFill(
            start_color=self.colors["total"],
            end_color=self.colors["total"],
            fill_type="solid"
        )
        tcell.alignment = self.align_center
        tcell.border = self.border
        
        mcell = ws.cell(row=row, column=7, value=total)
        mcell.font = Font(bold=True, size=13)
        mcell.fill = PatternFill(
            start_color=self.colors["total"],
            end_color=self.colors["total"],
            fill_type="solid"
        )
        mcell.alignment = self.align_right
        mcell.border = self.border
        self._currency(mcell)
        
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, row-2)}"
        self._set_col_widths_auto(ws)
    
    # ============================================
    # 🚨 HOJA 4: MOROSOS Y VENCIDOS
    # ============================================
    
    def _sheet_morosos(self, estudiantes, pagos):
        """Hoja con estudiantes que requieren atención de cobranza"""
        ws = self.workbook.create_sheet("🚨 Morosos")
        self._title_block(ws, "MOROSOS Y VENCIDOS", merge_to_col=9)
        
        pagos_por_cliente = self._pagos_por_cliente(pagos)
        
        headers = [
            "Estudiante", "Cédula", "Curso", "Email", "Teléfono",
            "Días", "Último Pago", "Total Pagado", "Estado"
        ]
        self._apply_table_header(ws, 4, headers)
        
        row = 5
        for e in estudiantes:
            if not getattr(e, "activo", True):
                continue
            
            dias = getattr(e, "dias_restantes", None)
            sin_cob = getattr(e, "mensualidades_canceladas", 0) == 0
            
            # Determinar si entra en cobranza
            if sin_cob:
                estado = "Sin cobertura"
            elif dias is not None and dias < 0:
                estado = "Vencido"
            elif dias is not None and dias <= 7:
                estado = "Por vencer"
            else:
                continue  # No necesita cobranza
            
            pagos_cliente = pagos_por_cliente.get(getattr(e, "id", None), [])
            total_pagado = sum(_money(p.monto) for p in pagos_cliente)
            
            # Último pago
            if pagos_cliente:
                ultimo_pago = pagos_cliente[0]
                fecha_ultimo = getattr(ultimo_pago, "fecha_pago", None)
                ultimo = fecha_ultimo.strftime("%d/%m/%Y") if fecha_ultimo else "-"
            else:
                ultimo = "-"
            
            curso = getattr(e, "curso", None)
            
            # Nombre completo
            nombre_completo = getattr(e, "nombre_completo", "")
            if not nombre_completo:
                nombre = getattr(e, "nombre", "")
                apellido = getattr(e, "apellido", "")
                nombre_completo = f"{nombre} {apellido}".strip()
            
            values = [
                nombre_completo,
                getattr(e, "cedula", None) or "Sin registrar",
                getattr(curso, "nombre", None) if curso else "Sin curso",
                getattr(e, "email", ""),
                getattr(e, "telefono", None) or "-",
                dias if dias is not None else ("0" if sin_cob else "N/A"),
                ultimo,
                total_pagado,
                estado,
            ]
            
            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = self.border
                cell.alignment = self.align_left
                cell.font = self.font_normal
                
                if col == 6:  # Días
                    cell.alignment = self.align_center
                if col == 8:  # Total pagado
                    cell.alignment = self.align_right
                    self._currency(cell)
                if col == 9:  # Estado
                    cell.alignment = self.align_center
                    if estado in ("Vencido", "Sin cobertura"):
                        cell.fill = PatternFill(
                            start_color=self.colors["danger"],
                            end_color=self.colors["danger"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.fill = PatternFill(
                            start_color=self.colors["warn"],
                            end_color=self.colors["warn"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True)
            
            row += 1
        
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, row-1)}"
        self._set_col_widths_auto(ws)
    
    # ============================================
    # 📚 HOJA 5: ANÁLISIS POR CURSO
    # ============================================
    
    def _sheet_cursos(self, estudiantes, pagos, cursos):
        """Hoja con análisis de ingresos y proyección por curso"""
        ws = self.workbook.create_sheet("📚 Cursos")
        self._title_block(ws, "ANÁLISIS POR CURSO", merge_to_col=9)
        
        # Agrupar pagos por curso
        pagos_por_curso = defaultdict(float)
        for p in pagos:
            c = getattr(p, "cliente", None)
            curso = getattr(c, "curso", None) if c else None
            nombre = getattr(curso, "nombre", None) if curso else "Sin curso"
            pagos_por_curso[nombre] += _money(p.monto)
        
        # Contar estudiantes por curso
        estudiantes_activos = [e for e in estudiantes if getattr(e, "activo", True)]
        conteo = defaultdict(int)
        for e in estudiantes_activos:
            curso = getattr(e, "curso", None)
            nombre = getattr(curso, "nombre", None) if curso else "Sin curso"
            conteo[nombre] += 1
        
        headers = [
            "Curso", "Activo", "Estudiantes", "Precio/Mes",
            "Ingreso Esperado", "Ingreso Real", "Diferencia",
            "Inscripción", "Descripción"
        ]
        self._apply_table_header(ws, 4, headers)
        
        row = 5
        for curso in cursos:
            nombre = getattr(curso, "nombre", "")
            activos_curso = conteo.get(nombre, 0)
            precio = _money(getattr(curso, "precio_mensual", 0))
            esperado = activos_curso * precio
            real = pagos_por_curso.get(nombre, 0.0)
            diff = real - esperado
            
            values = [
                nombre,
                "Sí" if getattr(curso, "activo", True) else "No",
                activos_curso,
                precio,
                esperado,
                real,
                diff,
                _money(getattr(curso, "precio_inscripcion", 0)),
                getattr(curso, "descripcion", None) or "-",
            ]
            
            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = self.border
                cell.alignment = self.align_left
                cell.font = self.font_normal
                
                if col in (2, 3):  # Activo, Estudiantes
                    cell.alignment = self.align_center
                if col in (4, 5, 6, 7, 8):  # Montos
                    cell.alignment = self.align_right
                    self._currency(cell)
                
                # Colorear diferencia
                if col == 7:
                    if diff >= 0:
                        cell.fill = PatternFill(
                            start_color=self.colors["ok"],
                            end_color=self.colors["ok"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                    else:
                        cell.fill = PatternFill(
                            start_color=self.colors["warn"],
                            end_color=self.colors["warn"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True)
            
            row += 1
        
        # Agregar "Sin curso" si existe
        if "Sin curso" in conteo or "Sin curso" in pagos_por_curso:
            activos_sin = conteo.get("Sin curso", 0)
            real_sin = pagos_por_curso.get("Sin curso", 0.0)
            
            values = ["Sin curso", "-", activos_sin, 0, 0, real_sin, real_sin, 0, "-"]
            
            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = self.border
                cell.alignment = self.align_left
                cell.font = self.font_normal
                
                if col in (2, 3):
                    cell.alignment = self.align_center
                if col in (4, 5, 6, 7, 8):
                    cell.alignment = self.align_right
                    self._currency(cell)
            
            row += 1
        
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, row-1)}"
        self._set_col_widths_auto(ws)
    
    # ============================================
    # 📄 REPORTES INDIVIDUALES (COMPATIBILIDAD)
    # ============================================
    
    def generar_reporte_estudiantes(self, estudiantes):
        """
        Genera reporte solo de estudiantes (para compatibilidad con rutas antiguas)
        Internamente genera el reporte completo
        """
        # Obtener pagos vacíos para mantener estructura
        return self.generar_reporte_completo(estudiantes, [], [])
    
    def generar_reporte_pagos(self, pagos, fecha_inicio=None, fecha_fin=None):
        """
        Genera reporte solo de pagos (para compatibilidad con rutas antiguas)
        """
        self.workbook = openpyxl.Workbook()
        
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        ws = self.workbook.create_sheet("💳 Pagos", 0)
        
        # Título con rango de fechas si aplica
        subtitle = "Generado: " + datetime.now().strftime('%d/%m/%Y %H:%M')
        if fecha_inicio and fecha_fin:
            subtitle = f"Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
        elif fecha_inicio:
            subtitle = f"Desde: {fecha_inicio.strftime('%d/%m/%Y')}"
        elif fecha_fin:
            subtitle = f"Hasta: {fecha_fin.strftime('%d/%m/%Y')}"
        
        self._title_block(ws, "REPORTE DE PAGOS", subtitle, merge_to_col=10)
        
        # Usar la misma lógica de _sheet_pagos
        headers = [
            "ID", "Fecha", "Estudiante", "Cédula", "Curso",
            "Periodo", "Monto", "Método", "Referencia", "Notas"
        ]
        self._apply_table_header(ws, 4, headers)
        
        row = 5
        total = 0.0
        
        for p in pagos:
            c = getattr(p, "cliente", None)
            curso = getattr(c, "curso", None) if c else None
            
            monto = _money(getattr(p, "monto", 0))
            total += monto
            
            if c:
                nombre_completo = getattr(c, "nombre_completo", "")
                if not nombre_completo:
                    nombre = getattr(c, "nombre", "")
                    apellido = getattr(c, "apellido", "")
                    nombre_completo = f"{nombre} {apellido}".strip()
            else:
                nombre_completo = "N/A"
            
            values = [
                getattr(p, "id", ""),
                getattr(p, "fecha_pago", None),
                nombre_completo,
                getattr(c, "cedula", None) if c else None,
                getattr(curso, "nombre", None) if curso else "Sin curso",
                getattr(p, "periodo", None) or "-",
                monto,
                getattr(p, "metodo_pago", None) or "-",
                getattr(p, "referencia", None) or "-",
                getattr(p, "notas", None) or "-",
            ]
            
            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = self.border
                cell.alignment = self.align_left
                cell.font = self.font_normal
                
                if col == 1:
                    cell.alignment = self.align_center
                if col == 2 and v:
                    if isinstance(v, datetime):
                        cell.value = v.strftime("%d/%m/%Y %H:%M")
                    cell.alignment = self.align_center
                if col == 7:
                    cell.alignment = self.align_right
                    self._currency(cell)
            
            row += 1
        
        # Total
        row += 1
        ws.merge_cells(f"A{row}:F{row}")
        tcell = ws.cell(row=row, column=1, value="TOTAL")
        tcell.font = Font(bold=True, size=13)
        tcell.fill = PatternFill(
            start_color=self.colors["total"],
            end_color=self.colors["total"],
            fill_type="solid"
        )
        tcell.alignment = self.align_center
        tcell.border = self.border
        
        mcell = ws.cell(row=row, column=7, value=total)
        mcell.font = Font(bold=True, size=13)
        mcell.fill = PatternFill(
            start_color=self.colors["total"],
            end_color=self.colors["total"],
            fill_type="solid"
        )
        mcell.alignment = self.align_right
        mcell.border = self.border
        self._currency(mcell)
        
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, row-2)}"
        self._set_col_widths_auto(ws)
        
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output
    
    def generar_reporte_proximos_vencer(self, estudiantes):
        """
        Genera reporte de próximos a vencer (para compatibilidad)
        """
        self.workbook = openpyxl.Workbook()
        
        if "Sheet" in self.workbook.sheetnames:
            self.workbook.remove(self.workbook["Sheet"])
        
        ws = self.workbook.create_sheet("🚨 Próximos a Vencer", 0)
        self._title_block(ws, "PRÓXIMOS A VENCER", merge_to_col=10)
        
        headers = [
            "Estudiante", "Cédula", "Email", "Teléfono", "Curso",
            "Días Restantes", "Fecha Vencimiento", "Mensualidades",
            "Última Actualización", "Prioridad"
        ]
        self._apply_table_header(ws, 4, headers)
        
        # Ordenar por días restantes (más urgente primero)
        estudiantes_ordenados = sorted(
            estudiantes,
            key=lambda e: getattr(e, "dias_restantes", 999) if getattr(e, "dias_restantes", None) is not None else 999
        )
        
        row = 5
        for e in estudiantes_ordenados:
            dias = getattr(e, "dias_restantes", None)
            
            # Determinar prioridad
            if dias is None or dias < 0:
                prioridad = "N/A"
                color = None
            elif dias <= 3:
                prioridad = "URGENTE"
                color = "danger"
            elif dias <= 7:
                prioridad = "Alta"
                color = "warn"
            else:
                prioridad = "Media"
                color = "info"
            
            curso = getattr(e, "curso", None)
            nombre_completo = getattr(e, "nombre_completo", "")
            if not nombre_completo:
                nombre = getattr(e, "nombre", "")
                apellido = getattr(e, "apellido", "")
                nombre_completo = f"{nombre} {apellido}".strip()
            
            values = [
                nombre_completo,
                getattr(e, "cedula", None) or "Sin registrar",
                getattr(e, "email", ""),
                getattr(e, "telefono", None) or "-",
                getattr(curso, "nombre", None) if curso else "Sin curso",
                dias if dias is not None else "N/A",
                getattr(e, "fecha_fin", None),
                getattr(e, "mensualidades_canceladas", 0),
                datetime.now().strftime("%d/%m/%Y"),
                prioridad,
            ]
            
            for col, v in enumerate(values, start=1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = self.border
                cell.alignment = self.align_left
                cell.font = self.font_normal
                
                if col in (6, 8):  # Días, Mensualidades
                    cell.alignment = self.align_center
                if col == 7 and v:  # Fecha vencimiento
                    if isinstance(v, datetime):
                        cell.value = v.strftime("%d/%m/%Y")
                    cell.alignment = self.align_center
                if col == 10:  # Prioridad
                    cell.alignment = self.align_center
                    if color == "danger":
                        cell.fill = PatternFill(
                            start_color=self.colors["danger"],
                            end_color=self.colors["danger"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
                    elif color == "warn":
                        cell.fill = PatternFill(
                            start_color=self.colors["warn"],
                            end_color=self.colors["warn"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True)
                    elif color == "info":
                        cell.fill = PatternFill(
                            start_color=self.colors["info"],
                            end_color=self.colors["info"],
                            fill_type="solid"
                        )
                        cell.font = Font(bold=True, color="FFFFFF")
            
            row += 1
        
        ws.freeze_panes = "A5"
        ws.auto_filter.ref = f"A4:{get_column_letter(len(headers))}{max(4, row-1)}"
        self._set_col_widths_auto(ws)
        
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        return output


# ============================================
# INSTANCIA GLOBAL (como tu sistema actual)
# ============================================

excel_generator = ExcelReportGenerator()